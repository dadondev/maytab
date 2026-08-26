
from openpyxl import load_workbook
import re

# ERP-generated exports contain many sheets (html0, html0 (2), ...), each
# holding a distinct set of classes. We must scan every sheet to gather all schedules.
pattern = re.compile(r"^(1[0-2]|[1-9])-([ABVGDE])$")
short_names_of_days = {
    "Du": "monday",
    "Se": "tuesday",
    "Ch": "wednesday",
    "Pa": "thursday",
    "Ju": "friday",
    "Sh": "saturday",
}


def _text(value):
    """Return a trimmed str, or None for empty/None cells."""
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def _subject_count(table) -> int:
    return sum(len(subjects) for subjects in table.values())


def _parse_sheet(ws):
    """
    Parse a single worksheet into `{ (grade, letter): {day: [subjects...] } }`.
    The header row is detected by a 'Kun' marker; class columns are anchored
    by their exact positions, so an extra 'Vaqt' column does not shift them.
    """
    schedules = {}
    class_columns = []  # list of (column_index, (grade, letter))
    prev_day = None

    for row in ws.iter_rows(values_only=True):
        cells = list(row)
        day_cell = ""
        if cells and cells[0] is not None:
            day_cell = str(cells[0]).strip()

        # Detect the header row (first cell == "Kun").
        if day_cell == "Kun":
            class_columns = []
            for index, value in enumerate(cells):
                text = _text(value)
                if text:
                    match = pattern.match(text)
                    if match:
                        grade, letter = match.group(1), match.group(2)
                        class_columns.append((index, (grade, letter)))
            continue

        if not class_columns:
            continue

        # A day marker row (e.g. "Du"). Register every class subject for that day.
        if day_cell in short_names_of_days:
            prev_day = short_names_of_days[day_cell]
            for index, (grade, letter) in class_columns:
                if index < len(cells):
                    subject = _text(cells[index])
                    if subject:
                        schedules.setdefault((grade, letter), {}).setdefault(prev_day, []).append(subject)
            continue

        # Continuation row for the current day (period number present).
        order = cells[1] if len(cells) > 1 else None
        if prev_day and order is not None:
            for index, (grade, letter) in class_columns:
                if index < len(cells):
                    subject = _text(cells[index])
                    if subject:
                        schedules.setdefault((grade, letter), {}).setdefault(prev_day, []).append(subject)

    return schedules


def get_data(file: str):
    wb = load_workbook(file, read_only=True, data_only=True)

    # Merge schedules from all sheets. The same class may appear on several
    # sheets (e.g. a duplicate export); keep the most complete version.
    merged = {}
    for ws in wb.worksheets:
        sheet_schedules = _parse_sheet(ws)
        for key, table in sheet_schedules.items():
            if key not in merged:
                merged[key] = table
            elif _subject_count(table) > _subject_count(merged[key]):
                merged[key] = table

    wb.close()

    grades = sorted({key[0] for key in merged}, key=int)
    # Convert to the [{ "name", "grade", "table" }, ...] shape the DB expects.
    # `name` is the full class id (e.g. "1-A") so it is globally unique and
    # shows meaningfully in the UI (the DB imposes unique=True on Group.name).
    tables = [
        {"name": f"{grade}-{letter}", "grade": grade, "table": table}
        for (grade, letter), table in sorted(
            merged.items(), key=lambda item: (int(item[0][0]), item[0][1])
        )
    ]
    return [grades, tables]
